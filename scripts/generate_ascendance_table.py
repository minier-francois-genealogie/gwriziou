#!/usr/bin/env python3
"""Generate ascendance table from GEDCOM + sources/documents/ with explicit column names."""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from act_path_normalize import normalize_commune, normalize_given_full, normalize_surname
from analyze_ascendance import find_root, parse_line
from gedcom_dates import parse_gedcom_dates
from paths import ACTES_DIR, GEDCOM_PATH

OUT_XLSX = Path(__file__).resolve().parent.parent.parent / "data" / "ascendance_ged_actes.xlsx"

ACT_FILE_RE = re.compile(
    r"^(?P<nom>[^_]+)__(?P<prenoms>.+)__(?P<type>[NDMP])__"
    r"(?P<date>(?:\d{4}|XXXX)-(?:\d{2}|XX)-(?:\d{2}|XX))__"
    r"(?P<dept>\d+|XX)__(?P<commune>.+?)"
    r"(?:__(?P<suffix>[^.]+))?\.(?P<ext>\w+)$",
    re.I,
)

COLUMNS = [
    "Warning",
    "GED_Generation",
    "GED_Sosa",
    "GED_Nom",
    "Acte_Nom",
    "GED_Prenom",
    "Acte_Prenom",
    "GED_DateNaissance",
    "Acte_DateNaissance",
    "GED_LieuNaissance",
    "Acte_LieuNaissance",
    "GED_DateMariage",
    "Acte_DateMariage",
    "GED_LieuMariage",
    "Acte_LieuMariage",
    "GED_DateDeces",
    "Acte_DateDeces",
    "GED_LieuDeces",
    "Acte_LieuDeces",
]

# Paires GED / Acte à comparer (colonnes côte à côte)
COMPARE_PAIRS = [
    ("GED_Nom", "Acte_Nom", "nom"),
    ("GED_Prenom", "Acte_Prenom", "prenom"),
    ("GED_DateNaissance", "Acte_DateNaissance", "date"),
    ("GED_LieuNaissance", "Acte_LieuNaissance", "lieu"),
    ("GED_DateMariage", "Acte_DateMariage", "date"),
    ("GED_LieuMariage", "Acte_LieuMariage", "lieu"),
    ("GED_DateDeces", "Acte_DateDeces", "date"),
    ("GED_LieuDeces", "Acte_LieuDeces", "lieu"),
]


@dataclass
class ActRecord:
    nom: str
    prenoms: str
    type: str
    date_iso: str
    dept: str
    commune: str


@dataclass
class ActBundle:
    nom: str = ""
    prenom: str = ""
    naissance: ActRecord | None = None
    mariage: ActRecord | None = None
    deces: ActRecord | None = None


def _flush_person_note(person: dict) -> None:
    pending = person.get("_pending_note")
    if pending:
        person.setdefault("notes", []).append(" ".join(pending))
        person["_pending_note"] = None


def _finalize_person(person: dict) -> dict:
    _flush_person_note(person)
    person.pop("_pending_note", None)
    return person


def load_gedcom(path: Path) -> tuple[dict[str, dict], dict[str, dict]]:
    individuals: dict[str, dict] = {}
    families: dict[str, dict] = {}
    cur_i: dict | None = None
    cur_f: dict | None = None
    ctx: str | None = None
    name_set = False

    with path.open(encoding="utf-8-sig", errors="replace") as fh:
        for line in fh:
            line = line.rstrip("\r\n")
            p = parse_line(line)
            if not p:
                continue
            level, tag, val = p

            if level == 0 and tag.startswith("@"):
                if val == "INDI":
                    if cur_i:
                        individuals[cur_i["id"]] = _finalize_person(cur_i)
                    cur_i = {
                        "id": tag,
                        "given": "",
                        "surn": "",
                        "sex": "",
                        "famc": None,
                        "fams": [],
                        "birt": None,
                        "birt_plac": None,
                        "deat": None,
                        "deat_plac": None,
                        "occu": None,
                        "nick": [],
                        "notes": [],
                        "_pending_note": None,
                    }
                    cur_f = None
                    ctx = None
                    name_set = False
                elif val == "FAM":
                    if cur_i:
                        individuals[cur_i["id"]] = _finalize_person(cur_i)
                        cur_i = None
                    if cur_f:
                        families[cur_f["id"]] = cur_f
                    cur_f = {
                        "id": tag,
                        "husb": None,
                        "wife": None,
                        "chil": [],
                        "marr": None,
                        "marr_plac": None,
                    }
                    ctx = None
                continue

            if cur_i:
                if level == 1 and tag == "NAME" and not name_set:
                    m = re.match(r"([^/]*)/([^/]*)/?", val)
                    if m:
                        cur_i["given"] = m.group(1).strip()
                        cur_i["surn"] = m.group(2).strip()
                        name_set = True
                    ctx = None
                elif level == 1 and tag == "SURN":
                    cur_i["surn"] = val.strip()
                elif level == 1 and tag == "GIVN":
                    cur_i["given"] = val.strip()
                elif level == 1 and tag == "SEX":
                    cur_i["sex"] = val.strip()
                elif level == 1 and tag == "FAMC":
                    cur_i["famc"] = val.strip()
                elif level == 1 and tag == "FAMS":
                    cur_i["fams"].append(val.strip())
                elif level == 1 and tag == "BIRT":
                    ctx = "BIRT"
                elif level == 1 and tag == "DEAT":
                    ctx = "DEAT"
                elif level == 1 and tag == "OCCU" and val.strip():
                    if not cur_i.get("occu"):
                        cur_i["occu"] = val.strip().strip('"')
                elif level == 1 and tag == "NOTE":
                    _flush_person_note(cur_i)
                    cur_i["_pending_note"] = [val.strip()]
                    ctx = None
                elif level == 1:
                    ctx = None
                elif level == 2 and tag in ("CONT", "CONC") and cur_i.get("_pending_note"):
                    cur_i["_pending_note"].append(val.strip())
                elif level == 2 and tag == "NICK":
                    cur_i.setdefault("nick", []).append(val.strip())
                elif level == 2 and ctx == "BIRT" and tag == "DATE":
                    cur_i["birt"] = val.strip()
                elif level == 2 and ctx == "BIRT" and tag == "PLAC":
                    cur_i["birt_plac"] = val.strip()
                elif level == 2 and ctx == "DEAT" and tag == "DATE":
                    cur_i["deat"] = val.strip()
                elif level == 2 and ctx == "DEAT" and tag == "PLAC":
                    cur_i["deat_plac"] = val.strip()
            elif cur_f:
                if level == 1 and tag == "HUSB":
                    cur_f["husb"] = val.strip()
                elif level == 1 and tag == "WIFE":
                    cur_f["wife"] = val.strip()
                elif level == 1 and tag == "CHIL":
                    cur_f["chil"].append(val.strip())
                elif level == 1 and tag == "MARR":
                    ctx = "MARR"
                elif level == 1:
                    ctx = None
                elif level == 2 and ctx == "MARR" and tag == "DATE":
                    cur_f["marr"] = val.strip()
                elif level == 2 and ctx == "MARR" and tag == "PLAC":
                    cur_f["marr_plac"] = val.strip()

    if cur_i:
        individuals[cur_i["id"]] = _finalize_person(cur_i)
    if cur_f:
        families[cur_f["id"]] = cur_f
    return individuals, families


def build_sosa_map(individuals: dict[str, dict], families: dict[str, dict], root_id: str) -> dict[int, str]:
    sosa_to_id: dict[int, str] = {1: root_id}
    stack = [1]
    seen: set[int] = set()
    while stack:
        s = stack.pop()
        if s in seen:
            continue
        seen.add(s)
        pid = sosa_to_id[s]
        if pid not in individuals:
            continue
        famc = individuals[pid].get("famc")
        if not famc or famc not in families:
            continue
        fam = families[famc]
        if fam.get("husb") and fam["husb"] in individuals:
            sosa_to_id[s * 2] = fam["husb"]
            stack.append(s * 2)
        if fam.get("wife") and fam["wife"] in individuals:
            sosa_to_id[s * 2 + 1] = fam["wife"]
            stack.append(s * 2 + 1)
    return sosa_to_id


def iso_to_fr(iso: str | None) -> str:
    if not iso:
        return ""
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", iso)
    if not m:
        y = re.search(r"(\d{4})", iso)
        return y.group(1) if y else iso
    y, mo, d = m.groups()
    if mo == "XX" and d == "XX":
        return y
    if mo == "XX":
        return f"{int(d):02d}/??/{y}"
    if d == "XX":
        return f"??/{int(mo):02d}/{y}"
    return f"{int(d):02d}/{int(mo):02d}/{y}"


def gedcom_date_fr(raw: str | None) -> str:
    return iso_to_fr(parse_gedcom_dates(raw)) if raw else ""


ARRONDISSEMENT_RE = re.compile(
    r"\s+\d{1,2}(?:er|ère|e|ème)?(?:\s+arrondissement)?$",
    re.I,
)


def ville_only(name: str) -> str:
    """Commune sans arrondissement (ex. « Paris 17e » → « Paris »)."""
    if not name:
        return ""
    s = name.strip()
    s = ARRONDISSEMENT_RE.sub("", s)
    return s.strip()


def gedcom_lieu(plac: str | None) -> str:
    """Première partie du PLAC GEDCOM = nom de la commune."""
    if not plac:
        return ""
    return ville_only(plac.split(",")[0].strip())


def act_lieu(dept: str, commune: str) -> str:
    return ville_only(commune.replace("_", " "))


def act_date_fr(date_iso: str) -> str:
    return iso_to_fr(date_iso)


def act_prenom_display(prenoms: str) -> str:
    return prenoms.replace("_", " ")


def norm_tokens(given: str) -> tuple[str, ...]:
    normalized = normalize_given_full(given.replace(",", " ").replace("-", " "))
    return tuple(t.upper() for t in normalized.split("_") if t)


def index_actes() -> dict[tuple[str, tuple[str, ...]], list[ActRecord]]:
    index: dict[tuple[str, tuple[str, ...]], list[ActRecord]] = {}
    for f in ACTES_DIR.rglob("*"):
        if not f.is_file() or f.suffix.lower() not in {".jpg", ".jpeg", ".pdf", ".png"}:
            continue
        if f.stat().st_size == 0:
            continue
        m = ACT_FILE_RE.match(f.name)
        if not m:
            continue
        d = m.groupdict()
        nom = normalize_surname(d["nom"])
        prenoms = normalize_given_full(d["prenoms"].replace("_", " "))
        tokens = tuple(t.upper() for t in prenoms.split("_") if t)
        key = (nom, tokens)
        rec = ActRecord(
            nom=d["nom"],
            prenoms=d["prenoms"],
            type=d["type"].upper(),
            date_iso=d["date"],
            dept=d["dept"],
            commune=d["commune"],
        )
        index.setdefault(key, []).append(rec)
    return index


def pick_first(records: list[ActRecord], act_type: str) -> ActRecord | None:
    typed = [r for r in records if r.type == act_type]
    if not typed:
        return None
    return sorted(typed, key=lambda r: r.date_iso)[0]


def match_score(tokens: tuple[str, ...], toks: tuple[str, ...]) -> int:
    if tokens == toks:
        return 1000
    if len(tokens) <= len(toks) and toks[: len(tokens)] == tokens:
        return 500 + len(tokens)
    if len(toks) <= len(tokens) and tokens[: len(toks)] == toks:
        return 400 + len(toks)
    common = 0
    for a, b in zip(tokens, toks):
        if a == b:
            common += 1
        else:
            break
    return common


def birth_year(iso: str | None) -> str | None:
    if not iso or len(iso) < 4:
        return None
    y = iso[:4]
    return y if y.isdigit() else None


def find_act_bundle(
    given: str,
    surn: str,
    index: dict[tuple[str, tuple[str, ...]], list[ActRecord]],
    gedcom_birth_iso: str | None = None,
) -> ActBundle:
    surn_n = normalize_surname(surn)
    tokens = norm_tokens(given)
    if not surn_n or not tokens:
        return ActBundle()

    scored: list[tuple[int, tuple[str, tuple[str, ...]]]] = []
    for key, records in index.items():
        if key[0] != surn_n:
            continue
        score = match_score(tokens, key[1])
        if score < 1:
            continue
        scored.append((score, key))

    scored.sort(key=lambda x: (-x[0], x[1]))
    ged_y = birth_year(parse_gedcom_dates(gedcom_birth_iso) if gedcom_birth_iso else None)

    for score, key in scored:
        records = index[key]
        n = pick_first(records, "N")
        m = pick_first(records, "M")
        d = pick_first(records, "D")

        if ged_y:
            act_y = birth_year(n.date_iso if n else None)
            if act_y and act_y != ged_y:
                continue

        ref = n or m or d
        if not ref:
            continue

        return ActBundle(
            nom=ref.nom,
            prenom=act_prenom_display(ref.prenoms),
            naissance=n,
            mariage=m,
            deces=d,
        )

    return ActBundle()


def first_marriage(person: dict, families: dict[str, dict]) -> tuple[str, str]:
    for fam_id in person.get("fams") or []:
        fam = families.get(fam_id)
        if not fam:
            continue
        if fam.get("marr") or fam.get("marr_plac"):
            return gedcom_date_fr(fam.get("marr")), gedcom_lieu(fam.get("marr_plac"))
    return "", ""


def act_row_part(rec: ActRecord | None) -> tuple[str, str]:
    if not rec:
        return "", ""
    return act_date_fr(rec.date_iso), act_lieu(rec.dept, rec.commune)


def has_acte_event(row: dict, date_col: str, lieu_col: str) -> bool:
    return bool(cell_text(row.get(date_col)) or cell_text(row.get(lieu_col)))


def missing_acte_warning(row: dict) -> str:
    if not cell_text(row.get("Acte_Nom")):
        missing = ["N", "M", "D"]
    else:
        missing = []
        if not has_acte_event(row, "Acte_DateNaissance", "Acte_LieuNaissance"):
            missing.append("N")
        if not has_acte_event(row, "Acte_DateMariage", "Acte_LieuMariage"):
            missing.append("M")
        if not has_acte_event(row, "Acte_DateDeces", "Acte_LieuDeces"):
            missing.append("D")
    if not missing:
        return ""
    return f"Manque Acte {'-'.join(missing)}"


def missing_ged_warning(row: dict) -> str:
    """Acte scanné mais ni date ni lieu côté GEDCOM pour cet événement."""
    missing: list[str] = []

    def ged_empty(date_col: str, lieu_col: str) -> bool:
        return not cell_text(row.get(date_col)) and not cell_text(row.get(lieu_col))

    if has_acte_event(row, "Acte_DateNaissance", "Acte_LieuNaissance") and ged_empty(
        "GED_DateNaissance", "GED_LieuNaissance"
    ):
        missing.append("N")
    if has_acte_event(row, "Acte_DateMariage", "Acte_LieuMariage") and ged_empty(
        "GED_DateMariage", "GED_LieuMariage"
    ):
        missing.append("M")
    if has_acte_event(row, "Acte_DateDeces", "Acte_LieuDeces") and ged_empty(
        "GED_DateDeces", "GED_LieuDeces"
    ):
        missing.append("D")
    if not missing:
        return ""
    return f"Manque GED {'-'.join(missing)}"


def row_has_ged_acte_diff(row: dict) -> bool:
    for ged_col, acte_col, kind in COMPARE_PAIRS:
        if values_differ(row.get(ged_col), row.get(acte_col), kind):
            return True
    return False


def build_warning_parts(row: dict) -> list[tuple[str, str]]:
    """Ordre : Manque GED (rouge), Diff GED/Acte (rouge), Manque Acte (orange)."""
    parts: list[tuple[str, str]] = []
    mg = missing_ged_warning(row)
    if mg:
        parts.append((mg, "red"))
    if row_has_ged_acte_diff(row):
        parts.append(("Diff GED/Acte", "red"))
    ma = missing_acte_warning(row)
    if ma:
        parts.append((ma, "orange"))
    return parts


def apply_warning_cell(cell, parts: list[tuple[str, str]]) -> None:
    """Texte multi-lignes ; une seule couleur par cellule (Excel stable, sans rich text)."""
    from openpyxl.styles import Font

    if not parts:
        cell.value = None
        return

    cell.value = "\n".join(text for text, _ in parts)
    has_red = any(color == "red" for _, color in parts)
    cell.font = Font(
        bold=True,
        color="FF0000" if has_red else "9C5700",
    )


def build_row(
    sosa: int,
    person: dict,
    families: dict[str, dict],
    act_index: dict[tuple[str, tuple[str, ...]], list[ActRecord]],
) -> dict[str, str]:
    gen = sosa.bit_length() - 1
    m_date, m_lieu = first_marriage(person, families)
    acts = find_act_bundle(person["given"], person["surn"], act_index, person.get("birt"))
    n_date, n_lieu = act_row_part(acts.naissance)
    a_m_date, a_m_lieu = act_row_part(acts.mariage)
    d_date, d_lieu = act_row_part(acts.deces)

    row = {
        "GED_Generation": gen,
        "GED_Sosa": sosa,
        "GED_Nom": person["surn"],
        "Acte_Nom": acts.nom,
        "GED_Prenom": person["given"],
        "Acte_Prenom": acts.prenom,
        "GED_DateNaissance": gedcom_date_fr(person.get("birt")),
        "Acte_DateNaissance": n_date,
        "GED_LieuNaissance": gedcom_lieu(person.get("birt_plac")),
        "Acte_LieuNaissance": n_lieu,
        "GED_DateMariage": m_date,
        "Acte_DateMariage": a_m_date,
        "GED_LieuMariage": m_lieu,
        "Acte_LieuMariage": a_m_lieu,
        "GED_DateDeces": gedcom_date_fr(person.get("deat")),
        "Acte_DateDeces": d_date,
        "GED_LieuDeces": gedcom_lieu(person.get("deat_plac")),
        "Acte_LieuDeces": d_lieu,
    }
    row["Warning"] = "\n".join(t for t, _ in build_warning_parts(row))
    return row


def cell_text(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def parse_fr_to_iso(s: str) -> str | None:
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    m = re.fullmatch(r"(\d{4})", s)
    return f"{m.group(1)}-01-01" if m else None


def norm_compare(val: str, kind: str) -> str:
    if not val:
        return ""
    if kind == "nom":
        return normalize_surname(val)
    if kind == "prenom":
        return normalize_given_full(val.replace(",", " ").replace("-", " "))
    if kind == "lieu":
        return normalize_commune(val)
    if kind == "date":
        iso = parse_fr_to_iso(val)
        return iso or val
    return val.upper()


def values_differ(ged_val, acte_val, kind: str) -> bool:
    g = cell_text(ged_val)
    a = cell_text(acte_val)
    if not g or not a:
        return False
    return norm_compare(g, kind) != norm_compare(a, kind)


def collect_mismatch_cells(df: pd.DataFrame) -> list[tuple[int, str]]:
    red: list[tuple[int, str]] = []
    for idx, row in df.iterrows():
        for ged_col, acte_col, kind in COMPARE_PAIRS:
            if values_differ(row[ged_col], row[acte_col], kind):
                red.append((idx, ged_col))
                red.append((idx, acte_col))
    return red


HEADER_ROW_COUNT = 3
DATA_START_ROW = 4

# Blocs événement : (colonne début, libellé) — colonne 1 = Alertes
EVENT_BLOCKS = [(8, "Naissance"), (12, "Mariage"), (16, "Décès")]


def fix_merged_cell_borders(ws, thin) -> None:
    """Bordures correctes sur plages fusionnées : dé-fusionner, contourner, re-fusionner."""
    from openpyxl.styles import Border, Side
    from openpyxl.utils.cell import range_boundaries

    no_side = Side(border_style=None)
    merges = [str(r) for r in ws.merged_cells.ranges]
    for ref in merges:
        ws.unmerge_cells(ref)

    for ref in merges:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=thin if col == min_col else no_side,
                    right=thin if col == max_col else no_side,
                    top=thin if row == min_row else no_side,
                    bottom=thin if row == max_row else no_side,
                )
        ws.merge_cells(ref)


def write_styled_excel(
    df: pd.DataFrame,
    path: Path,
    mismatch_cells: list[tuple[int, str]],
) -> int:
    from openpyxl import Workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Ascendance"
    ncol = len(COLUMNS)
    col_idx = {name: i + 1 for i, name in enumerate(COLUMNS)}

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="000000")
    all_borders = Border(left=thin, right=thin, top=thin, bottom=thin)
    black_fill = PatternFill("solid", fgColor="000000")
    red_font = Font(color="FF0000")
    sep_font = Font(color="FFFFFF", bold=True)

    gen_counts = df.groupby("GED_Generation").size().to_dict()

    def hdr(row: int, col: int, value: str) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = all_borders

    # Colonne Alertes (fusion verticale lignes 1-3)
    ws.merge_cells(start_row=1, start_column=1, end_row=HEADER_ROW_COUNT, end_column=1)
    hdr(1, 1, "Alertes")

    # Génération + SOSA : 3 niveaux (fusion verticale lignes 1-3)
    for c, label in ((2, "Génération"), (3, "SOSA")):
        ws.merge_cells(start_row=1, start_column=c, end_row=HEADER_ROW_COUNT, end_column=c)
        hdr(1, c, label)

    # Nom + Prénom : 2 niveaux (titre lignes 1-2, GED/Acte ligne 3)
    for start, title in ((4, "Nom"), (6, "Prénom")):
        ws.merge_cells(start_row=1, start_column=start, end_row=2, end_column=start + 1)
        hdr(1, start, title)
        hdr(3, start, "GED")
        hdr(3, start + 1, "Acte")

    # Naissance / Mariage / Décès : 3 niveaux (événement → Date/Lieu → GED/Acte)
    for start, title in EVENT_BLOCKS:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + 3)
        hdr(1, start, title)
        ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=start + 1)
        ws.merge_cells(start_row=2, start_column=start + 2, end_row=2, end_column=start + 3)
        hdr(2, start, "Date")
        hdr(2, start + 2, "Lieu")
        hdr(3, start, "GED")
        hdr(3, start + 1, "Acte")
        hdr(3, start + 2, "GED")
        hdr(3, start + 3, "Acte")

    for r in range(1, HEADER_ROW_COUNT + 1):
        ws.row_dimensions[r].height = 22
    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=1)

    # Données + séparateurs noirs entre générations
    excel_row = DATA_START_ROW
    df_index_to_excel: dict[int, int] = {}
    prev_gen = None
    separator_count = 0
    separator_rows: set[int] = set()

    for idx, row in df.iterrows():
        gen = row["GED_Generation"]
        if prev_gen is not None and gen != prev_gen:
            separator_rows.add(excel_row)
            count = gen_counts.get(prev_gen, 0)
            max_gen = 2**prev_gen
            label = f"Génération {prev_gen} : {count}/{max_gen}"
            ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=ncol)
            sep_cell = ws.cell(row=excel_row, column=1, value=label)
            sep_cell.fill = black_fill
            sep_cell.font = sep_font
            sep_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[excel_row].height = 16
            excel_row += 1
            separator_count += 1

        df_index_to_excel[idx] = excel_row
        for col_name in COLUMNS:
            val = row[col_name]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=excel_row, column=col_idx[col_name])
            if col_name == "Warning":
                apply_warning_cell(cell, build_warning_parts(row.to_dict()))
            else:
                cell.value = val
        prev_gen = gen
        excel_row += 1

    last_row = excel_row - 1
    data_align = Alignment(vertical="top", wrap_text=True)
    for r in range(1, last_row + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.border = all_borders
            if r in separator_rows:
                if not isinstance(cell, MergedCell):
                    cell.fill = black_fill
            elif r <= HEADER_ROW_COUNT:
                cell.alignment = header_align
                cell.fill = header_fill
            else:
                cell.alignment = data_align
                if c == 1:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    fix_merged_cell_borders(ws, thin)

    for df_idx, col_name in mismatch_cells:
        if df_idx not in df_index_to_excel or col_name not in col_idx:
            continue
        ws.cell(
            row=df_index_to_excel[df_idx],
            column=col_idx[col_name],
        ).font = red_font

    # Largeurs de colonnes
    widths = {
        1: 22,
        2: 10,
        3: 8,
    }
    for c in range(4, ncol + 1):
        widths[c] = 14
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return separator_count


def main() -> None:
    individuals, families = load_gedcom(GEDCOM_PATH)
    root_id = find_root(individuals)
    sosa_to_id = build_sosa_map(individuals, families, root_id)
    act_index = index_actes()

    rows: list[dict] = []
    for sosa in sorted(sosa_to_id):
        pid = sosa_to_id[sosa]
        person = individuals.get(pid)
        if not person:
            continue
        rows.append(build_row(sosa, person, families, act_index))

    df = pd.DataFrame(rows, columns=COLUMNS)
    mismatches = collect_mismatch_cells(df)
    separators = write_styled_excel(df, OUT_XLSX, mismatches)

    with_actes = df["Acte_Nom"].astype(bool).sum()
    print(f"Lignes SOSA : {len(df)}")
    print(f"Génération max : {df['GED_Generation'].max()}")
    print(f"Avec actes scannés : {with_actes}")
    print(f"Séparateurs de génération : {separators}")
    print(f"Cellules en écart (rouge) : {len(mismatches)}")
    print(f"Fichier : {OUT_XLSX}")


if __name__ == "__main__":
    main()

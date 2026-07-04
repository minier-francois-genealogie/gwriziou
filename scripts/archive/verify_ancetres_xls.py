#!/usr/bin/env python3
"""Verify and enrich ancetres.xls against GEDCOM and actes/."""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from act_path_normalize import normalize_commune, normalize_given_full, normalize_surname
from analyze_ascendance import analyze, find_root, load_gedcom, person_label

XLS_PATH = Path(r"C:\Projet\Perso\genealogie\data\ancetres.xls")
OUT_XLSX = Path(r"C:\Projet\Perso\genealogie\data\ancetres_corrige.xlsx")
ACTES_DIR = Path(r"C:\Projet\Perso\genealogie\data\actes")
REPORT_PATH = Path(__file__).resolve().parent / "ancetres_verification.json"

PERSON_FOLDER_RE = re.compile(
    r"^(?P<nom>[^_]+)__(?P<prenoms>.+)__(?P<bdate>(?:\d{4}|XXXX)-(?:\d{2}|XX)-(?:\d{2}|XX))__"
    r"(?P<dept>\d+|XX)__(?P<bcommune>.+)$"
)
ACT_FILE_RE = re.compile(
    r"^(?P<nom>[^_]+)__(?P<prenoms>.+)__(?P<type>[NDM])__",
    re.I,
)


def parse_fr_date(val) -> str | None:
    if pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    m = re.search(r"(\d{4})", s)
    return f"{m.group(1)}-01-01" if m else None


def norm_name(s: str | None) -> str:
    if not s or pd.isna(s):
        return ""
    return normalize_given_full(str(s).replace(",", " ").replace("-", " "))


def norm_surn(s: str | None) -> str:
    if not s or pd.isna(s):
        return ""
    return normalize_surname(str(s))


def gedcom_date_iso(raw: str | None) -> str | None:
    if not raw:
        return None
    from gedcom_dates import parse_gedcom_dates

    parsed = parse_gedcom_dates(raw)
    if parsed:
        return parsed
    m = re.search(r"(\d{4})", raw)
    return f"{m.group(1)}-01-01" if m else None


def commune_from_plac(plac: str | None) -> str:
    if not plac:
        return ""
    return normalize_commune(plac.split(",")[0])


def index_actes() -> dict[tuple[str, tuple[str, ...]], dict[str, int]]:
    """Map (surname, given tokens) -> {N: count, M: count, D: count}."""
    index: dict[tuple[str, tuple[str, ...]], dict[str, int]] = {}
    for f in ACTES_DIR.rglob("*"):
        if not f.is_file() or f.suffix.lower() not in {".jpg", ".jpeg", ".pdf", ".png"}:
            continue
        if f.stat().st_size == 0:
            continue
        m = ACT_FILE_RE.match(f.name)
        if not m:
            continue
        nom = m.group("nom").upper()
        prenoms = tuple(t.upper() for t in m.group("prenoms").split("_") if t)
        key = (nom, prenoms)
        t = m.group("type").upper()
        index.setdefault(key, {"N": 0, "M": 0, "D": 0})
        index[key][t] = index[key].get(t, 0) + 1
    return index


def find_act_counts(nom: str, prenom: str, act_index: dict) -> tuple[int, dict[str, int]]:
    surn = norm_surn(nom)
    tokens = tuple(t.upper() for t in norm_name(prenom).split("_") if t)
    if not surn or not tokens:
        return 0, {}
    # exact match
    key = (surn, tokens)
    if key in act_index:
        d = act_index[key]
        return sum(d.values()), d
    # prefix match on given tokens
    for (s, toks), d in act_index.items():
        if s == surn and toks[: len(tokens)] == tokens[: len(toks)]:
            return sum(d.values()), d
        if s == surn and tokens[: len(toks)] == toks:
            return sum(d.values()), d
    return 0, {}


def iso_to_fr_date(iso: str | None) -> str | None:
    if not iso:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", iso)
    if not m:
        y = re.search(r"(\d{4})", iso)
        return f"01/01/{y.group(1)}" if y else None
    y, mo, d = m.groups()
    if mo == "XX" or d == "XX":
        return f"{y}" if mo == "XX" else f"{int(d):02d}/{int(mo):02d}/{y}"
    return f"{int(d):02d}/{int(mo):02d}/{y}"


def first_marriage_date(person: dict, families: dict[str, dict]) -> str | None:
    for fam_id in person.get("fams") or []:
        fam = families.get(fam_id)
        if fam and fam.get("marr"):
            return gedcom_date_iso(fam["marr"])
    return None


def marriage_date_fr(person: dict, families: dict[str, dict]) -> str | None:
    g_m = first_marriage_date(person, families)
    return iso_to_fr_date(g_m) if g_m else None


def is_female_person(person: dict, sosa: int) -> bool:
    sex = (person.get("sex") or "").upper()
    if sex == "F":
        return True
    if sex == "M":
        return False
    # Ahnentafel : sosa 2n+1 = mère (sauf racine sosa 1)
    return sosa > 1 and sosa % 2 == 1


def birth_from_acte(row, act_breakdown: dict[str, int], marriage_fr: str | None) -> str | None:
    """Colonne Acte = date d'acte de naissance ; ne pas y lire un acte de mariage."""
    if pd.isna(row["Acte"]):
        return None
    src = str(row["Acte"]).strip()
    if not src or src.lower() == "nan":
        return None
    if marriage_fr and src == marriage_fr:
        return None
    if act_breakdown.get("M", 0) > 0 and act_breakdown.get("N", 0) == 0:
        return None
    if act_breakdown.get("N", 0) > 0 or not any(act_breakdown.values()):
        return src
    return None


def apply_red_cells(path: Path, cells: list[tuple[int, str]], columns: list[str]) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    if not cells:
        return
    wb = load_workbook(path)
    ws = wb["GED sheet"]
    col_idx = {name: i + 1 for i, name in enumerate(columns)}
    red = Font(color="FF0000")
    for row_idx, col_name in cells:
        if col_name not in col_idx:
            continue
        # pandas row 0 -> excel row 2 (1 header line)
        cell = ws.cell(row=row_idx + 2, column=col_idx[col_name])
        cell.font = red
    wb.save(path)


def collapse_ref(sosa: int, sosa_to_id: dict[int, str], individuals: dict) -> int | None:
    """Find earlier SOSA with same GEDCOM id (pedigree collapse)."""
    pid = sosa_to_id.get(sosa)
    if not pid:
        return None
    for s in sorted(s for s in sosa_to_id if s < sosa):
        if sosa_to_id[s] == pid:
            return s
    return None


def main() -> None:
    individuals, families = load_gedcom(XLS_PATH.parent / "ged" / "fminier.ged")
    root_id = find_root(individuals)
    ged_report = analyze(individuals, families, root_id)
    sosa_to_id = {1: root_id}
    stack = [1]
    seen = set()
    while stack:
        s = stack.pop()
        if s in seen:
            continue
        seen.add(s)
        pid = sosa_to_id[s]
        if pid not in individuals:
            continue
        famc = individuals[pid].get("famc")
        if not famc or famc not in families:
            continue
        fam = families[famc]
        if fam.get("husb") and fam["husb"] in individuals:
            fs = s * 2
            sosa_to_id[fs] = fam["husb"]
            stack.append(fs)
        if fam.get("wife") and fam["wife"] in individuals:
            ms = s * 2 + 1
            sosa_to_id[ms] = fam["wife"]
            stack.append(ms)

    df = pd.read_excel(XLS_PATH, sheet_name=0, header=1)
    act_index = index_actes()

    issues: list[dict] = []
    red_cells: list[tuple[int, str]] = []
    df["Ref collapse"] = pd.Series([""] * len(df), dtype=object)
    df["Ecart GEDCOM"] = pd.Series([""] * len(df), dtype=object)
    df["Actes N"] = pd.NA
    df["Actes M"] = pd.NA
    df["Actes D"] = pd.NA
    dates_filled = {"naissance": 0, "mariage": 0, "deces": 0}

    for idx, row in df.iterrows():
        sosa = int(row["#Sosa"])
        if pd.isna(row["Nom"]):
            ref = collapse_ref(sosa, sosa_to_id, individuals)
            if ref:
                df.at[idx, "Ref collapse"] = f"= sosa_{ref}"
            else:
                issues.append({"sosa": sosa, "type": "EMPTY_ROW_NO_REF"})
            continue

        pid = sosa_to_id.get(sosa)
        if not pid:
            issues.append({"sosa": sosa, "type": "SOSA_NOT_IN_GEDCOM_TREE", "name": f"{row['Nom']} {row['Prenom']}"})
            df.at[idx, "Ecart GEDCOM"] = "hors arbre GEDCOM"
            continue

        gp = individuals[pid]
        ecarts: list[str] = []
        _, act_breakdown = find_act_counts(str(row["Nom"]), str(row["Prenom"]), act_index)
        marriage_fr = marriage_date_fr(gp, families)
        female = is_female_person(gp, sosa)

        # Compléter dates manquantes (seront en rouge)
        if pd.isna(row["Date"]):
            src = None
            g_b = gedcom_date_iso(gp.get("birt"))
            if g_b:
                candidate = iso_to_fr_date(g_b)
                if candidate and candidate != marriage_fr:
                    src = candidate
            if not src:
                src = birth_from_acte(row, act_breakdown, marriage_fr)
            if src:
                df.at[idx, "Date"] = src
                red_cells.append((idx, "Date"))
                dates_filled["naissance"] += 1
                row = df.loc[idx]

        # Heredis : date de mariage sur la fiche du mari, pas de la femme
        if pd.isna(row["Date.1"]) and not female:
            if marriage_fr:
                df.at[idx, "Date.1"] = marriage_fr
                red_cells.append((idx, "Date.1"))
                dates_filled["mariage"] += 1
                row = df.loc[idx]

        if pd.isna(row["Date.2"]) and gp.get("has_deat"):
            g_d = gedcom_date_iso(gp.get("deat"))
            if g_d:
                df.at[idx, "Date.2"] = iso_to_fr_date(g_d)
                red_cells.append((idx, "Date.2"))
                dates_filled["deces"] += 1
                row = df.loc[idx]

        if norm_surn(row["Nom"]) != norm_surn(gp["surn"]):
            ecarts.append(f"nom: {row['Nom']} vs {gp['surn']}")
        if norm_name(row["Prenom"]) != norm_name(gp["given"]):
            # allow partial: first names match
            x_t = norm_name(row["Prenom"]).split("_")
            g_t = norm_name(gp["given"]).split("_")
            if x_t[:2] != g_t[:2] and x_t[0] != g_t[0]:
                ecarts.append(f"prénom: {row['Prenom']} vs {gp['given']}")

        x_b = parse_fr_date(row["Date"])
        g_b = gedcom_date_iso(gp.get("birt"))
        if x_b and g_b and x_b[:4] != g_b[:4]:
            ecarts.append(f"naissance: {row['Date']} vs {gp.get('birt')}")
        if not pd.isna(row["Date"]) and not g_b and (idx, "Date") not in red_cells:
            ecarts.append("naissance absente du GEDCOM")

        x_d = parse_fr_date(row["Date.2"])
        g_d = gedcom_date_iso(gp.get("deat"))
        if x_d and g_d and x_d[:4] != g_d[:4]:
            ecarts.append(f"décès: {row['Date.2']} vs {gp.get('deat')}")

        if ecarts:
            df.at[idx, "Ecart GEDCOM"] = "; ".join(ecarts)
            issues.append({"sosa": sosa, "type": "GEDCOM_MISMATCH", "details": ecarts, "name": person_label(gp)})

        total = sum(act_breakdown.values())
        df.at[idx, "Nombre d'actes"] = total
        df.at[idx, "Actes N"] = act_breakdown.get("N", 0)
        df.at[idx, "Actes M"] = act_breakdown.get("M", 0)
        df.at[idx, "Actes D"] = act_breakdown.get("D", 0)

        # Flag acte column vs filesystem
        if pd.notna(row["Acte"]) and act_breakdown.get("N", 0) == 0:
            issues.append({"sosa": sosa, "type": "XLS_ACTE_N_BUT_NO_FILE", "name": person_label(gp)})
        if pd.notna(row["Acte.1"]) and act_breakdown.get("M", 0) == 0:
            issues.append({"sosa": sosa, "type": "XLS_ACTE_M_BUT_NO_FILE", "name": person_label(gp)})
        if pd.notna(row["Acte.2"]) and act_breakdown.get("D", 0) == 0:
            issues.append({"sosa": sosa, "type": "XLS_ACTE_D_BUT_NO_FILE", "name": person_label(gp)})

    # Summary stats
    filled = df[df["Nom"].notna()]
    report = {
        "at": datetime.now().isoformat(),
        "source": str(XLS_PATH),
        "output": str(OUT_XLSX),
        "rows_total": len(df),
        "rows_filled": len(filled),
        "rows_collapse_empty": int(df["Nom"].isna().sum()),
        "gedcom_mismatches": int((df["Ecart GEDCOM"] != "").sum()),
        "with_actes": int((filled["Nombre d'actes"].fillna(0) > 0).sum()),
        "dates_filled_red": dates_filled,
        "dates_filled_total": sum(dates_filled.values()),
        "issue_count": len(issues),
        "issues_by_type": {},
        "issues_sample": issues[:50],
    }
    from collections import Counter

    report["issues_by_type"] = dict(Counter(i["type"] for i in issues))

    REPORT_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="GED sheet", index=False)
        issues_df = pd.DataFrame(issues)
        if not issues_df.empty:
            issues_df.to_excel(writer, sheet_name="Ecarts", index=False)

    apply_red_cells(OUT_XLSX, red_cells, list(df.columns))

    print(f"Lignes: {len(df)} ({len(filled)} renseignées, {report['rows_collapse_empty']} collapse vides)")
    print(f"Dates complétées (rouge): {report['dates_filled_total']} — {dates_filled}")
    print(f"Écarts GEDCOM: {report['gedcom_mismatches']}")
    print(f"Avec actes scannés: {report['with_actes']}")
    print(f"Problèmes: {report['issue_count']}")
    print("Par type:", report["issues_by_type"])
    print(f"Rapport: {REPORT_PATH}")
    print(f"Fichier corrigé: {OUT_XLSX}")


if __name__ == "__main__":
    main()
